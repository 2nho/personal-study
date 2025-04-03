import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# CSV 파일 로드
csv_file = 'data-1743579611789.csv'
csv_data = pd.read_csv(csv_file)

# 새 엑셀 워크북 생성
wb = Workbook()

# 스타일 설정
header_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# 테이블 이름 목록 추출
table_names = csv_data['table_name'].unique()

# 각 테이블 이름에 대해 새로운 시트 생성 및 데이터 삽입
for table_name in table_names:
    # 해당 테이블 데이터 필터링
    table_data = csv_data[csv_data['table_name'] == table_name]
    
    # 새로운 시트 생성
    sheet = wb.create_sheet(title=table_name[:31])  # 시트 이름은 31자로 제한
    
    # 1번 행: 테이블정의서 (셀 병합 및 가운데 정렬)
    sheet.merge_cells('A1:I1')
    sheet['A1'] = '테이블정의서'
    sheet['A1'].font = header_font
    #sheet['A1'].alignment = center_alignment
    sheet['A1'].fill = gray_fill
    
    # 2번 행: 주제영역명 / 작성일 / 작성자
    #sheet.append(['주제영역명', '', '', '작성일', '2025-04-03', '', '작성자', '김인호'])
    sheet['A2'] = '주제영역명'
    sheet['D2'] = '작성일'
    sheet['G2'] = '작성자'
    sheet.merge_cells('B2:C2')
    sheet.merge_cells('E2:F2')
    sheet['E2'] = '2025-04-03'
    sheet.merge_cells('H2:I2')
    sheet['H2'] = '김인호'
    for col in ['A2', 'D2', 'G2']:
        sheet[col].fill = gray_fill
        sheet[col].font = header_font
    
    # 3번 행: 테이블ID / 테이블명
    #sheet.append(['테이블ID', '', table_name, '테이블명', '', table_name, '', '', ''])
    sheet['A3'] = '테이블ID'
    sheet['D3'] = '테이블명'
    sheet.merge_cells('B3:C3')
    sheet['B3'] = table_name
    sheet.merge_cells('E3:I3')
    sheet['E3'] = table_name # 일단 같게 처리
    for col in ['A3', 'D3']:
        sheet[col].fill = gray_fill
        sheet[col].font = header_font
    
    # 4번 행: 테이블 설명
    #sheet.append(['테이블설명', '', f'{table_name} 테이블', '', '', '', '', '', ''])
    sheet['A4'] = '테이블설명'
    sheet.merge_cells('B4:I4')
    sheet['B4'] = f'{table_name} 테이블'
    sheet['A4'].fill = gray_fill
    sheet['A4'].font = header_font
    
    # 5번 행: 헤더
    headers = ["No.", "컬럼ID", "컬럼명", "타입", "길이", "NULL", "KEY", "DEFAULT", "비고"]
    sheet.append(headers)
    
    # 헤더 스타일 적용
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=5, column=col_num)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = gray_fill
        #cell.border = thin_border
    
    # 데이터 삽입 (No.는 테이블별로 1부터 시작)
    for idx, row in enumerate(table_data.iterrows(), start=1):
        row_data = row[1]  # row[0]은 인덱스, row[1]은 실제 데이터
        sheet.append([
            idx,  # No. (테이블별로 1부터 시작)
            row_data['column_name'],  # 컬럼ID
            row_data['column_name'],  # 컬럼명 (같은 값 사용)
            row_data['type'],  # 타입
            row_data['length'],  # 길이
            row_data['is_nullable'],  # NULL 여부
            row_data['pk'],  # KEY 여부
            row_data['column_default'],  # DEFAULT 값
            row_data['comment']  # 비고
        ])
        
        # 데이터 행에 테두리 적용
        # for col_num in range(1, 10):
        #     sheet.cell(row=idx+5, column=col_num).border = thin_border
    
    # 인덱스 및 업무규칙 행 추가
    next_row = len(table_data) + 6  # 헤더행(5) + 데이터 행 수 + 1
    # 변수설정

    # 인덱스 행 추가 및 스타일 적용
    sheet.merge_cells(start_row=next_row, start_column=2, end_row=next_row, end_column=3)  # B~C 병합
    sheet.cell(row=next_row, column=1, value="인 덱 스")
    sheet.cell(row=next_row, column=1).fill = gray_fill
    sheet.cell(row=next_row, column=1).font = header_font

   
    sheet.merge_cells(start_row=next_row, start_column=5, end_row=next_row, end_column=9)
    sheet.cell(row=next_row, column=4, value="인 덱 스 키")
    sheet.cell(row=next_row, column=4).fill = gray_fill
    sheet.cell(row=next_row, column=4).font = header_font
    
    # 업무규칙 행 추가 및 스타일 적용
    sheet.merge_cells(start_row=next_row+1, start_column=2, end_row=next_row+1, end_column=9)  # B~I 병합
    sheet.cell(row=next_row+1, column=1, value="업무규칙")
    sheet.cell(row=next_row+1, column=1).fill = gray_fill
    sheet.cell(row=next_row+1, column=1).font = header_font
    
    
    # 인덱스와 업무규칙 행에 테두리 적용
    # for row_num in [next_row, next_row+1]:
    #     for col_num in range(1, 10): # 1~9(A~I)
    #         sheet.cell(row=row_num, column=col_num).border = thin_border
    
    #  사용중인 쉘 전부 테두리 처리
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # 열 너비 조정
    sheet.column_dimensions['A'].width = 10  # No.
    sheet.column_dimensions['B'].width = 15  # 컬럼ID
    sheet.column_dimensions['C'].width = 20  # 컬럼명
    sheet.column_dimensions['D'].width = 15  # 타입
    sheet.column_dimensions['E'].width = 10  # 길이
    sheet.column_dimensions['F'].width = 10  # NULL
    sheet.column_dimensions['G'].width = 10  # KEY
    sheet.column_dimensions['H'].width = 15  # DEFAULT
    sheet.column_dimensions['I'].width = 20  # 비고

# 기본 시트 삭제
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# 엑셀 파일 저장
output_file = '테스트엑셀.xlsx'
wb.save(output_file)
print(f"엑셀 파일이 성공적으로 생성되었습니다: {output_file}")
